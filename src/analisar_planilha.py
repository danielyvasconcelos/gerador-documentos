import openpyxl
import os

def analisar_planilha(caminho_arquivo):
    """Analisa a estrutura da planilha RPONTES"""
    try:
        wb = openpyxl.load_workbook(caminho_arquivo)
        
        print(f"Arquivo: {os.path.basename(caminho_arquivo)}")
        print(f"Abas disponíveis: {wb.sheetnames}")
        print("-" * 50)
        
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            print(f"\nAba: '{nome_aba}'")
            
            # Mostra cabeçalhos (primeira linha)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            print(f"Colunas ({len(headers)}): {headers}")
            
            # Conta linhas com dados
            linhas_dados = 0
            for row in ws.iter_rows(min_row=2):
                if any(cell.value for cell in row):
                    linhas_dados += 1
                else:
                    break
                    
            print(f"Linhas de dados: {linhas_dados}")
            
            # Mostra exemplo da primeira linha de dados
            if linhas_dados > 0:
                primeira_linha = [cell.value for cell in ws[2]]
                print(f"Exemplo: {primeira_linha[:3]}...")  # Primeiros 3 valores
                
        wb.close()
        
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {caminho_arquivo}")
        print("Certifique-se de que o arquivo está na pasta 'data/'")
    except Exception as e:
        print(f"Erro ao analisar: {e}")

if __name__ == "__main__":
    # Tenta diferentes extensões
    caminhos = [
        "../data/RPONTES.xlsx",
        "../data/RPONTES.xls",
        "../data/rpontes.xlsx"
    ]
    
    for caminho in caminhos:
        if os.path.exists(caminho):
            analisar_planilha(caminho)
            break
    else:
        print("Arquivo RPONTES não encontrado na pasta data/")
        print("Coloque o arquivo na pasta: data/RPONTES.xlsx")